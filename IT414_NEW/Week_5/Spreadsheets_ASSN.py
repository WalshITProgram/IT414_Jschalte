import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from gspread import authorize as gspread_authorize, WorksheetNotFound
from google.oauth2.service_account import Credentials

# Function to scrape data from the provided URL
def scrape_sales_data(url):
    # Send a GET request to the URL
    response = requests.get(url)
    # Parse the HTML content of the page
    soup = BeautifulSoup(response.text, 'html.parser')
    # Find the table containing sales data
    table = soup.find('table')
    # Initialize dictionary to store product and quantity
    data = {'Product': [], 'Quantity': []}
    # Iterate over each row in the table
    for row in table.find_all('tr')[1:]:
        cells = row.find_all('td')
        product = cells[3].text.strip()
        quantity = int(cells[4].text.strip())
        # If product already exists in data, update quantity
        if product in data['Product']:
            index = data['Product'].index(product)
            data['Quantity'][index] += quantity
        # If product is new, add it to data
        else:
            data['Product'].append(product)
            data['Quantity'].append(quantity)
    return data

# Function to calculate the average number of items sold per order
def calculate_average(data):
    total_items = sum(data['Quantity'])
    num_orders = len(data['Quantity'])
    return total_items / num_orders

# Function to generate Excel spreadsheet
def generate_excel(data, average):
    # Create a new Workbook
    wb = Workbook()
    # Activate the default worksheet
    ws = wb.active
    # Set title of worksheet
    ws.title = "Sales Data"
    # Append column headers
    ws.append(['Product', 'Quantity'])
    # Append product and quantity data
    for product, quantity in zip(data['Product'], data['Quantity']):
        ws.append([product, quantity])
    # Append average calculation
    ws.append(['Average Number of Items Sold per Order', average])
    # Format title cell
    title_cell = ws['A1']
    title_cell.value = "Sales Data"
    title_cell.font = Font(bold=True, color="0000FF")  # Blue-colored font
    # Format column headers
    title_row = ws[1]
    for cell in title_row:
        cell.font = Font(bold=True)
    # Center-align all cells
    for row in ws.iter_rows(min_row=2, max_row=len(data['Product'])+1, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True)
    # Save the workbook
    wb.save('sales_data.xlsx')

# Function to authorize access to Google Sheets
def authorize(credentials_file):
    credentials = Credentials.from_service_account_file(credentials_file)
    gc = gspread_authorize(credentials)
    return gc

# Function to generate Google Sheet
def generate_google_sheet(data, average, credentials_file):
    # Authorize access to Google Sheets
    gc = authorize(credentials_file)
    try:
        # Try to open existing sheet named "SalesData"
        sheet = gc.open("SalesData")
    except WorksheetNotFound:
        # If sheet does not exist, create a new one
        sheet = gc.create("SalesData")
    # Get the first worksheet in the spreadsheet
    worksheet = sheet.get_worksheet(0)
    # Update column headers
    worksheet.update('A1', 'Product')
    worksheet.update('B1', 'Quantity')
    # Update product and quantity data
    for i, (product, quantity) in enumerate(zip(data['Product'], data['Quantity']), start=2):
        worksheet.update('A{}'.format(i), product)
        worksheet.update('B{}'.format(i), quantity)
    # Update average calculation
    worksheet.update('A{}'.format(len(data['Product'])+2), 'Average Number of Items Sold per Order')
    worksheet.update('B{}'.format(len(data['Product'])+2), average)

# Main function
def main():
    # URL of the sales data report
    url = "https://ool-content.walshcollege.edu/CourseFiles/IT/IT414/MASTER/Week05/WI20-Assignment/sales_data.html"
    # Scrape sales data from the URL
    data = scrape_sales_data(url)
    # Calculate the average number of items sold per order
    average = calculate_average(data)
    # Path to the Google Sheets credentials file
    credentials_file = r'path_to_your_credentials_file'  # Update with the path to your Google Sheets credentials file
    
    # Generate Excel spreadsheet
    generate_excel(data, average)
    # Generate Google Sheet
    generate_google_sheet(data, average, credentials_file)

    print("Excel spreadsheet and Google Sheet generated successfully.")

if __name__ == "__main__":
    main()

